from __future__ import annotations

import base64
import csv
import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import UUID, uuid4

from danfer_os.models.importer import (
    FileUpload,
    ImportConfiguration,
    ImportField,
    ImportHistory,
    ImportPreview,
    ImportProfile,
    ImportProfileCreate,
    ImportValidation,
    ValidationIssue,
)
from danfer_os.services.technical_library import TechnicalLibrary


class ImporterError(ValueError):
    pass


@dataclass
class _Session:
    filename: str
    columns: list[str]
    rows: list[list[str]]


class ImporterService:
    def __init__(self, library: TechnicalLibrary) -> None:
        self._library = library
        self._sessions: dict[UUID, _Session] = {}
        self._profiles: dict[UUID, ImportProfile] = {}
        self._history: list[ImportHistory] = []

    def preview(self, upload: FileUpload) -> ImportPreview:
        try:
            content = base64.b64decode(upload.content_base64, validate=True)
        except ValueError as error:
            raise ImporterError("conteúdo base64 inválido") from error
        suffix = Path(upload.filename).suffix.casefold()
        if suffix == ".csv":
            rows, sheets = self._read_csv(content, upload.delimiter)
        elif suffix == ".xlsx":
            rows, sheets = self._read_xlsx(content, upload.sheet)
        elif suffix == ".xls":
            rows, sheets = self._read_xls(content, upload.sheet)
        else:
            raise ImporterError("formato aceito: .csv, .xlsx ou .xls")
        header_index = upload.header_row - 1
        if header_index >= len(rows):
            raise ImporterError("linha de cabeçalho fora do arquivo")
        columns = self._unique_columns(rows[header_index])
        data_rows = [
            (row + [""] * len(columns))[:len(columns)]
            for row in rows[header_index + 1:]
            if any(cell.strip() for cell in row)
        ]
        session_id = uuid4()
        self._sessions[session_id] = _Session(upload.filename, columns, data_rows)
        return ImportPreview(
            session_id=session_id,
            filename=upload.filename,
            columns=columns,
            rows=data_rows[:20],
            total_rows=len(data_rows),
            sheets=sheets,
        )

    @staticmethod
    def _read_csv(content: bytes, delimiter: str | None) -> tuple[list[list[str]], list[str]]:
        text = content.decode("utf-8-sig")
        dialect = None
        if delimiter is None:
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
            except csv.Error:
                delimiter = ";"
        reader = csv.reader(io.StringIO(text), dialect=dialect) if dialect else csv.reader(
            io.StringIO(text), delimiter=delimiter or ";"
        )
        return [[str(cell).strip() for cell in row] for row in reader], []

    @staticmethod
    def _read_xlsx(content: bytes, sheet: str | None) -> tuple[list[list[str]], list[str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise ImporterError("dependência openpyxl não instalada") from error
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = workbook.sheetnames
        if sheet and sheet not in sheets:
            raise ImporterError("aba não encontrada")
        worksheet = workbook[sheet or sheets[0]]
        rows = [["" if value is None else str(value) for value in row] for row in worksheet.values]
        workbook.close()
        return rows, sheets

    @staticmethod
    def _read_xls(content: bytes, sheet: str | None) -> tuple[list[list[str]], list[str]]:
        try:
            import xlrd
        except ImportError as error:
            raise ImporterError("dependência xlrd não instalada") from error
        workbook = xlrd.open_workbook(file_contents=content)
        sheets = workbook.sheet_names()
        if sheet and sheet not in sheets:
            raise ImporterError("aba não encontrada")
        worksheet = workbook.sheet_by_name(sheet or sheets[0])
        rows = [
            [str(worksheet.cell_value(row, column)).strip() for column in range(worksheet.ncols)]
            for row in range(worksheet.nrows)
        ]
        return rows, sheets

    @staticmethod
    def _unique_columns(values: list[str]) -> list[str]:
        columns, counts = [], {}
        for index, value in enumerate(values, start=1):
            base = value.strip() or f"coluna_{index}"
            counts[base] = counts.get(base, 0) + 1
            columns.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
        return columns

    def validate(self, session_id: UUID, config: ImportConfiguration) -> ImportValidation:
        session = self._sessions.get(session_id)
        if session is None:
            raise LookupError(session_id)
        source_indexes = {column: index for index, column in enumerate(session.columns)}
        for mapping in config.mappings:
            if mapping.source_column not in source_indexes:
                raise ImporterError(f"coluna não encontrada: {mapping.source_column}")
        catalog = self._library.list()
        materials = {item.material.casefold() for item in catalog if item.material}
        codes = {
            code.casefold()
            for item in catalog
            for code in (item.danfer_code, item.customer_code)
            if code
        }
        normalized, issues = [], []
        for row_number, cells in enumerate(session.rows, start=2):
            item: dict[str, str | float] = dict(config.fixed_values)
            for mapping in config.mappings:
                if mapping.target_field != ImportField.IGNORE:
                    item[mapping.target_field.value] = cells[source_indexes[mapping.source_column]]
            self._validate_row(row_number, item, codes, materials, issues)
            normalized.append(item)
        invalid = len({issue.row for issue in issues if issue.critical})
        return ImportValidation(
            session_id=session_id,
            normalized_rows=normalized,
            issues=issues,
            valid_rows=len(normalized) - invalid,
            invalid_rows=invalid,
        )

    @staticmethod
    def _validate_row(
        row: int,
        item: dict[str, str | float],
        codes: set[str],
        materials: set[str],
        issues: list[ValidationIssue],
    ) -> None:
        code = str(item.get("codigo_danfer") or item.get("codigo_cliente") or "").strip()
        if not code:
            issues.append(ValidationIssue(row=row, field="codigo", message="código vazio", critical=True))
        elif codes and code.casefold() not in codes:
            issues.append(ValidationIssue(row=row, field="codigo", message="código não localizado", critical=False))
        try:
            quantity = float(str(item.get("quantidade", "")).replace(",", "."))
            if quantity <= 0:
                raise ValueError
            item["quantidade"] = quantity
        except ValueError:
            issues.append(ValidationIssue(row=row, field="quantidade", message="quantidade inválida", critical=True))
        material = str(item.get("material", "")).strip()
        if material and materials and material.casefold() not in materials:
            issues.append(ValidationIssue(row=row, field="material", message="material desconhecido", critical=False))

    def create_profile(self, data: ImportProfileCreate) -> ImportProfile:
        profile = ImportProfile(**data.model_dump())
        self._profiles[profile.id] = profile
        return profile.model_copy(deep=True)

    def list_profiles(self, customer: str | None = None) -> list[ImportProfile]:
        profiles = self._profiles.values()
        if customer:
            profiles = (item for item in profiles if item.customer.casefold() == customer.casefold())
        return [item.model_copy(deep=True) for item in profiles]

    def finish(
        self,
        validation: ImportValidation,
        customer: str = "",
        profile_id: UUID | None = None,
    ) -> ImportHistory:
        session = self._sessions.get(validation.session_id)
        if session is None:
            raise LookupError(validation.session_id)
        history = ImportHistory(
            session_id=validation.session_id,
            filename=session.filename,
            customer=customer,
            profile_id=profile_id,
            total_rows=len(session.rows),
            valid_rows=validation.valid_rows,
            invalid_rows=validation.invalid_rows,
        )
        self._history.append(history)
        return history

    def history(self) -> list[ImportHistory]:
        return [item.model_copy(deep=True) for item in self._history]
